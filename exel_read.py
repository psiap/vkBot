from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Let's use the hello_world spreadsheet since it has less data
workbook = load_workbook(filename="test.xlsx")
sheet = workbook['E40']
logo = Image("python.png")

# A bit of resizing to not fill the whole spreadsheet with the logo
logo.height = 100
logo.width = 100

sheet.add_image(logo, "B7")

workbook.save(filename="test2.xlsx")


import openpyxl
import zipfile
from shutil import copyfile
from shutil import rmtree
import os

PAD = os.getcwd()

wb = openpyxl.load_workbook('1.xlsm')
sheet = wb['E40']
logo = Image("logo.png")

logo.height = 100
logo.width = 100

sheet.add_image(logo, "B7")
sheet.cell(row=2, column=3).value = 'Edited'   # example
#####

wb.save('2.xlsx')


with zipfile.ZipFile('1.xlsm', 'r') as z:
    z.extractall('./xlsm/')

with zipfile.ZipFile('2.xlsx', 'r') as z:
    z.extractall('./xlsx/')

copyfile('./xlsm/[Content_Types].xml','./xlsx/[Content_Types].xml')
copyfile('./xlsm/xl/_rels/workbook.xml.rels','./xlsx/xl/_rels/workbook.xml.rels')
#copyfile('./xlsm/xl/vbaProject.bin','./xlsx/xl/vbaProject.bin')

z = zipfile.ZipFile('2.zip', 'w')

os.chdir('./xlsx')

for root, dirs, files in os.walk('./'):
        for file in files:
            z.write(os.path.join(root, file))
z.close()

#clean
os.chdir(PAD)
rmtree('./xlsm/')
rmtree('./xlsx/')
os.remove('./2.xlsx')
os.rename('2.zip', '2.xlsm')