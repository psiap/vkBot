import time

from bs4 import BeautifulSoup
import requests
import pandas
import openpyxl as ox
from openpyxl import load_workbook


def update_spreadsheet(path: str, _df, starcol: int = 1, startrow: int = 1, sheet_name: str = "ToUpdate"):
    '''

    :param path: Путь до файла Excel
    :param _df: Датафрейм Pandas для записи
    :param starcol: Стартовая колонка в таблице листа Excel, куда буду писать данные
    :param startrow: Стартовая строка в таблице листа Excel, куда буду писать данные
    :param sheet_name: Имя листа в таблице Excel, куда буду писать данные
    :return:
    '''
    wb = ox.load_workbook(path)
    for ir in range(0, len(_df)):
        for ic in range(0, len(_df.iloc[ir])):
            wb[sheet_name].cell(startrow + ir, starcol + ic).value = _df.iloc[ir][ic]


    wb.save(path)

#Получение названия:сылок
def read_excel_links_name():
    titles_and_references_array = {}
    excel_data_df = pandas.ExcelFile('price.xlsx')
    #print(excel_data_df.sheet_names)
    for i in excel_data_df.sheet_names:
        try:
            #print(pandas.read_excel('price.xlsx', sheet_name=i).iloc[5, 1])
            titles_and_references_array[i] = pandas.read_excel('price.xlsx', sheet_name=i).iloc[5, 1]
        except Exception:
            pass
    return titles_and_references_array


def getting_product_links(url):
    s = requests.Session()
    s.headers.update({
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.9; rv:45.0) Gecko/20100101 Firefox/45.0'
        })
    request = s.get(url)

    if request.status_code != 200:
        print(f'ERROR: return False: {str(request.status_code)}')
        return False

    filteredNews = []

    soup = BeautifulSoup(request.text, "html.parser")

    allNews = soup.findAll('div', class_='product-name')
    for data in allNews:
        if data.find('a').get('href') is not None:
            filteredNews.append("https://inled.ru/" + data.find('a').get('href'))

    temp_count_replace = 1
    while True:
        url_temp = f"{url}/p/{str(temp_count_replace)}"
        request = s.get(url_temp)
        if request.status_code != 200:
            #print(f'ERROR: return False: {str(request.status_code)}')
            break
        soup = BeautifulSoup(request.text, "html.parser")

        allNews = soup.findAll('div', class_='product-name')
        for data in allNews:
            if data.find('a').get('href') is not None:
                filteredNews.append("https://inled.ru/" + data.find('a').get('href'))
        temp_count_replace += 1

    return filteredNews

def taking_the_attributes(links_url):
    s = requests.Session()
    s.headers.update({
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.9; rv:45.0) Gecko/20100101 Firefox/45.0'
    })
    row = 8
    for url in links_url:
        request = s.get(url)
        if request.status_code != 200:
            print(f'ERROR: break: {str(request.status_code)}')
            break

        soup = BeautifulSoup(request.text, "html.parser")
        soup.find('div', class_='shop2-product-params')
        attributes_all_dict = {}
        #attributes_all = []
        #Получение артикула
        attributes_all_dict["артикул"] = [int(x) for x in soup.find('div', class_='shop2-product-article').text.split() if x.isdigit()]

        #Получение картинки
        attributes_all_dict["фото"] = soup.find('div', class_='card-slider__image').find('a').get('href')

        #Получение цены
        #attributes_all_dict["Цена"] = [int(x) for x in soup.find('div', class_='product-price').text.split() if x.isdigit()]
        attributes_all_dict["Цена"] = soup.find('div', class_='product-price').text
        #Получения названия
        attributes_all_dict["Наименование"] = soup.find('h1', class_='product-name').text

        #Получения сылки
        attributes_all_dict["url"] = url

        attributes_all_dict["RCI"] = '80'

        attributes_all_dict["Коэфф. мощности"] = '0,9'

        for art in soup.find('div', class_='shop2-product-params'):

            #мощность
            if 'Потребляемая мощность (Вт)' in art.text:
                attributes_all_dict["мощность"] = ''.join([str(x) for x in art.text if x.isdigit()])

            elif 'Световой поток (Лм)' in art.text:
                attributes_all_dict["световой поток"] = art.text.split(')')[1]
            elif 'Габаритные размеры' in art.text:
                attributes_all_dict["габариты,mm"] = art.text.replace('Габаритные размеры', '').replace(' мм', '')
            elif 'Вес' in art.text:
                attributes_all_dict["вес, g"] = art.text.replace('Вес', '').replace(' г.', '')
            elif 'Степень защиты (IP)' in art.text:
                attributes_all_dict["Степень защиты(IP)"] = art.text.split(')')[1]
            elif 'Угол рассеивания (Град.)' in art.text:
                attributes_all_dict["Угол светового пучка (градусы)"] = art.text.split(')')[1]
            elif 'Гарантийный срок (Мес.)' in art.text:
                attributes_all_dict["Гарантия"] = art.text.split(')')[1]
            elif 'Срок эксплуатации (ч.)' in art.text:
                attributes_all_dict["Срок эксплуатации, ч"] = art.text.split(')')[1]

            #attributes_all.append(art.text)
        #print(attributes_all)
        for x,i in enumerate(attributes_all_dict):
            print(f"{str(x)}){i} - {attributes_all_dict[i]}")
        df = pandas.DataFrame.from_dict(attributes_all_dict)
        update_spreadsheet('test.xlsx', df, 1, row,'E27')
        row += 1

        print('Норм')
        #time.sleep(50)

def final_list():
    #Получаем названия и ссылки
    mass_dict_links = read_excel_links_name()
    for key in mass_dict_links:
        #Получаем ссылки страницы
        links_url = getting_product_links(mass_dict_links[key])
        #Получаем атрибуты
        taking_the_attributes(links_url)
        #time.sleep(20)

final_list()

#1) Получаем ссылки всех товаров на странице
#links_url = getting_product_links('https://inled.ru/magazin/folder/svetilniki-zhkh')
#2) Получаем все атрибуты на странице
#taking_the_attributes(links_url)